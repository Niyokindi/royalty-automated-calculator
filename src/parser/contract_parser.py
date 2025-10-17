"""
Music Contract Royalty Parser using GPT-4
Extracts parties, works, and royalty shares from music contracts
Supports PDF and Excel files

Installation:
pip install openai PyMuPDF python-dotenv openpyxl

Setup:
1. Create a .env file with your OpenAI API key:
   OPENAI_API_KEY=sk-your-api-key-here

2. Place your contract file in the same directory or provide the path
"""

import os
import json
import re
from typing import Dict, List, Optional, Union
from dataclasses import dataclass, asdict
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
from openai import OpenAI
from dotenv import load_dotenv
import streamlit as st  # Safe to import here

# Load environment variables
load_dotenv()


@dataclass
class Party:
    """Represents a party involved in the contract"""
    name: str
    role: str
    additional_info: Optional[str] = None


@dataclass
class Work:
    """Represents a musical work"""
    title: str
    work_type: str  # Song, Album, EP, Single, etc.
    additional_info: Optional[str] = None


@dataclass
class RoyaltyShare:
    """Represents a royalty share distribution"""
    party_name: str
    royalty_type: str  # mechanical, publishing, performance, master, sync
    percentage: float
    terms: Optional[str] = None


@dataclass
class ContractData:
    """Complete extracted contract data"""
    parties: List[Party]
    works: List[Work]
    royalty_shares: List[RoyaltyShare]
    contract_summary: Optional[str] = None
    raw_text: Optional[str] = None


class MusicContractParser:
    #Parser for music contracts using GPT-4
    def __init__(self, api_key: Optional[str] = None):
        """
        Initialize the parser with OpenAI API key.
        Works in both local and Streamlit Cloud environments.
        """

        

        # Load API key from: parameter → secrets → env
        self.api_key = (
            api_key
            or st.secrets.get("OPENAI_API_KEY")
            or os.getenv("OPENAI_API_KEY")
        )
        st.write("✅ Loaded API key?", bool(self.api_key))

        if not self.api_key:
            raise ValueError("❌ OpenAI API key not found. Add it to Streamlit secrets or .env file.")

        self.client = OpenAI(api_key=self.api_key)
        self.model = "gpt-4o"  # Using GPT-4 Omni for better performance
    
    
    def extract_text_from_file(self, file_path: str) -> str:
        """
        Extract text content from a PDF or Excel file
        
        Args:
            file_path: Path to the file
            
        Returns:
            Extracted text content
        """
        file_extension = Path(file_path).suffix.lower()
        
        if file_extension == '.pdf':
            return self._extract_text_from_pdf(file_path)
        elif file_extension in ['.xlsx', '.xls']:
            return self._extract_text_from_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_extension}. Supported types: .pdf, .xlsx, .xls")
    
    def _extract_text_from_pdf(self, pdf_path: str) -> str:
        #Extract text content from a PDF file
        text = ""
        try:
            # Open the PDF
            doc = fitz.open(pdf_path)
            
            # Extract text from each page
            for page_num in range(len(doc)):
                page = doc[page_num]
                text += page.get_text()
                text += "\n\n"  # Add spacing between pages
            
            # Close the document
            doc.close()
            
            if not text.strip():
                raise Exception("No text could be extracted. PDF might be scanned images.")
                
        except Exception as e:
            raise Exception(f"Error extracting text from PDF: {str(e)}")
        
        return text
    
    def _extract_text_from_excel(self, excel_path: str) -> str:
        #Extract text content from an Excel file"
        text = ""
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n=== SHEET: {sheet_name} ===\n\n"
                
                # Get all rows
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    # Filter out None values and convert to strings
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    # Only add non-empty rows
                    if any(val.strip() for val in row_values):
                        text += " | ".join(row_values) + "\n"
            
            workbook.close()
            
        except Exception as e:
            raise Exception(f"Error extracting text from Excel: {str(e)}")
        
        return text
    
    def parse_contract(self, file_path: str) -> ContractData:
        """
        Parse a music contract file and extract all relevant information
        
        Args:
            file_path: Path to the contract file (PDF or Excel)
            
        Returns:
            ContractData object with all extracted information
        """
        print(f"Extracting text from file: {file_path}")
        contract_text = self.extract_text_from_file(file_path)
        
        if not contract_text.strip():
            raise ValueError("No text could be extracted from the file")
        
        print(f"Extracted {len(contract_text)} characters")
        print("Analyzing contract with GPT-4...")
        
        # Extract parties
        parties = self._extract_parties(contract_text)
        print(f"✓ Found {len(parties)} parties")
        
        # Extract works
        works = self._extract_works(contract_text)
        print(f"✓ Found {len(works)} works")
        
        # Extract royalty shares
        royalty_shares = self._extract_royalty_shares(contract_text)
        print(f"✓ Found {len(royalty_shares)} royalty share entries")
        
        # Generate summary
        summary = self._generate_summary(contract_text)
        print("✓ Generated contract summary")
        
        return ContractData(
            parties=parties,
            works=works,
            royalty_shares=royalty_shares,
            contract_summary=summary,
            raw_text=contract_text[:1000] + "..." if len(contract_text) > 1000 else contract_text
        )
    
    def _extract_parties(self, contract_text: str) -> List[Party]:
        #Extract all parties involved in the contract
        
        # Limit text size for API call
        text_sample = contract_text[:12000]
        
        prompt = f"""
        You are a music contract analyst. Extract all parties (people and entities) mentioned in this contract or document.
        
        For each party, identify:
        1. Full name
        2. Role (artist, producer, songwriter, composer, publisher, manager, label, etc.)
        3. Any additional relevant information
        
        Contract text:
        {text_sample}
        
        Return ONLY a JSON object with a "parties" array:
        {{
            "parties": [
                {{"name": "John Doe", "role": "artist", "additional_info": "performing artist"}},
                {{"name": "Jane Smith", "role": "producer", "additional_info": "executive producer"}}
            ]
        }}
        
        Be thorough but avoid duplicates. If uncertain about a role, use "party" as the role.
        """
        
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are a precise legal document analyst specializing in music contracts. Always return valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            parties_data = result.get('parties', [])
            
            parties = [
                Party(
                    name=p.get('name', ''),
                    role=p.get('role', 'party'),
                    additional_info=p.get('additional_info')
                )
                for p in parties_data if p.get('name')
            ]
            
            return parties
            
        except Exception as e:
            print(f"Warning: Error extracting parties: {str(e)}")
            return []
    
    def _extract_works(self, contract_text: str) -> List[Work]:
        #Extract all musical works mentioned in the contract
        
        text_sample = contract_text[:12000]
        
        prompt = f"""
        You are a music contract analyst. Extract all musical works mentioned in this contract or document.
        
        For each work, identify:
        1. Title of the work
        2. Type (Song, Album, EP, Single, Composition, Recording, etc.)
        3. Any additional relevant information
        
        Contract text:
        {text_sample}
        
        Return ONLY a JSON object with a "works" array:
        {{
            "works": [
                {{"title": "Song Name", "work_type": "Song", "additional_info": "lead single"}},
                {{"title": "Album Name", "work_type": "Album", "additional_info": "debut album"}}
            ]
        }}
        
        Be thorough but avoid duplicates.
        """
        
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are a precise legal document analyst specializing in music contracts. Always return valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            works_data = result.get('works', [])
            
            works = [
                Work(
                    title=w.get('title', ''),
                    work_type=w.get('work_type', 'Work'),
                    additional_info=w.get('additional_info')
                )
                for w in works_data if w.get('title')
            ]
            
            return works
            
        except Exception as e:
            print(f"Warning: Error extracting works: {str(e)}")
            return []
    
    def _extract_royalty_shares(self, contract_text: str) -> List[RoyaltyShare]:
        #Extract royalty share distributions from the contract
        
        text_sample = contract_text[:15000]
        
        prompt = f"""
        You are a music contract analyst. Extract ALL royalty share distributions from this contract or document.
        
        CRITICAL: Look for ALL mentions of percentage splits, revenue sharing, or royalty distributions throughout the entire text.
        This includes splits mentioned in streaming revenue
        
        For EACH party and EACH type of royalty mentioned, create a separate entry.
        
        For each royalty distribution, identify:
        1. Party name receiving the royalty (use exact name from contract)
        2. Type of royalty:  streaming
        3. Percentage share (as a decimal number, e.g., 30.0 for 30%)
        4. Any relevant terms or conditions
        
        Contract text:
        {text_sample}
        
        Return ONLY a JSON object with a "royalty_shares" array. Create ONE entry per party per royalty type:
        {{
            "royalty_shares": [
                {{"party_name": "John Doe", "royalty_type": "publishing", "percentage": 30.0, "terms": "split equally"}},
                {{"party_name": "John Doe", "royalty_type": "streaming", "percentage": 30.0, "terms": "net revenue from streaming"}},
                {{"party_name": "Karen Smith", "royalty_type": "publishing", "percentage": 30.0, "terms": "split equally"}},
                {{"party_name": "Karen Smith", "royalty_type": "streaming", "percentage": 30.0, "terms": "net revenue from streaming"}}
            ]
        }}
        
        IMPORTANT: 
        - If the same percentage applies to multiple royalty types, create separate entries for each type
        - Use the party names exactly as they appear in the contract
        - Be thorough - extract every single royalty split mentioned
        - Streaming revenue should be categorized as "streaming" type
        - Publishing splits should be categorized as "publishing" type
        - Net revenue can be "net_revenue" type or the specific type it relates to
        """
        
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are a precise legal document analyst specializing in music royalty contracts. Always return valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            
            result = json.loads(response.choices[0].message.content)
            royalties_data = result.get('royalty_shares', [])
            
            royalty_shares = [
                RoyaltyShare(
                    party_name=r.get('party_name', ''),
                    royalty_type=r.get('royalty_type', 'other'),
                    percentage=float(r.get('percentage', 0)),
                    terms=r.get('terms')
                )
                for r in royalties_data if r.get('party_name') and r.get('percentage') is not None
            ]
            
            return royalty_shares
            
        except Exception as e:
            print(f"Warning: Error extracting royalty shares: {str(e)}")
            return []
    
    def _generate_summary(self, contract_text: str) -> str:
        #Generate a concise summary of the contract
        
        text_sample = contract_text[:10000]
        
        prompt = f"""
        Provide a concise 3-4 sentence summary of this music contract or document, focusing on:
        - Who the main parties are
        - What the contract is about (which works)
        - Key financial/royalty terms
        
        Contract text:
        {text_sample}
        
        Return ONLY the summary text, no additional formatting.
        """
        
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are a concise legal document summarizer."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=300
            )
            
            return response.choices[0].message.content.strip()
            
        except Exception as e:
            print(f"Warning: Error generating summary: {str(e)}")
            return "Summary generation failed"
    
    def save_to_json(self, contract_data: ContractData, output_path: str):
        """
        Save extracted contract data to a JSON file
        
        Args:
            contract_data: The extracted contract data
            output_path: Path to save the JSON file
        """
        data_dict = {
            'parties': [asdict(p) for p in contract_data.parties],
            'works': [asdict(w) for w in contract_data.works],
            'royalty_shares': [asdict(r) for r in contract_data.royalty_shares],
            'contract_summary': contract_data.contract_summary
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data_dict, f, indent=2, ensure_ascii=False)
        
        print(f"Data saved to {output_path}")
    
    def print_summary(self, contract_data: ContractData):
        """Print a formatted summary of the extracted data"""
        
        print("\n" + "="*80)
        print("CONTRACT ANALYSIS SUMMARY")
        print("="*80)
        
        if contract_data.contract_summary:
            print(f"\nSummary:\n{contract_data.contract_summary}")
        
        print(f"\n👥 PARTIES ({len(contract_data.parties)}):")
        print("-" * 80)
        if contract_data.parties:
            for party in contract_data.parties:
                info = f" - {party.additional_info}" if party.additional_info else ""
                print(f"  • {party.name} ({party.role}){info}")
        else:
            print("  No parties extracted")
        
        print(f"\nWORKS ({len(contract_data.works)}):")
        print("-" * 80)
        if contract_data.works:
            for work in contract_data.works:
                info = f" - {work.additional_info}" if work.additional_info else ""
                print(f"  • {work.title} ({work.work_type}){info}")
        else:
            print("No works extracted")
        
        print(f"\nROYALTY SHARES ({len(contract_data.royalty_shares)}):")
        print("-" * 80)
        if contract_data.royalty_shares:
            for share in contract_data.royalty_shares:
                terms = f" - {share.terms}" if share.terms else ""
                print(f"  • {share.party_name}: {share.percentage}% ({share.royalty_type}){terms}")
        else:
            print("  No royalty shares extracted")
        
        print("\n" + "="*80)


def main():
    """Example usage"""
    
    # Initialize parser
    parser = MusicContractParser()
    
    # Path to your contract file (PDF or Excel)
    file_path = "/Users/kenjiniyokindi/Documents/GREENBOX ANALYTICS/Personal Projects/Music Data Projects/Royalty automation/Contracts/Scenario 2 ' Home' - Romes_Lebron Contract.pdf"
    
    try:
        # Parse the contract
        contract_data = parser.parse_contract(file_path)
        
        # Print summary
        parser.print_summary(contract_data)
        
        # Save to JSON
        output_path = "contract_analysis.json"
        parser.save_to_json(contract_data, output_path)
        
    except Exception as e:
        print(f"\nError: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()