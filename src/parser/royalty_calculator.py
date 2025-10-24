"""
Royalty Payment Calculator
Calculates payments from royalty statements and music contracts

Installation:
pip install openpyxl

Usage:
    from royalty_calculator import RoyaltyCalculator
    from contract_parser import MusicContractParser
    
    calculator = RoyaltyCalculator()
    payments = calculator.calculate_payments("contract.pdf", "statement.xlsx")
"""
import openpyxl
from typing import Dict, List, Optional
from dataclasses import dataclass, asdict
import os
from openai import OpenAI
from dotenv import load_dotenv
import streamlit as st  # Safe to import here

# Load environment variables
load_dotenv()

# Import the contract parser
from parser.contract_parser import MusicContractParser, ContractData, Party, Work, RoyaltyShare
@dataclass
class RoyaltyPayment:
    #Represents a calculated royalty payment
    song_title: str
    party_name: str
    role: str
    royalty_type: str
    percentage: float
    total_royalty: float
    amount_to_pay: float


class RoyaltyCalculator:
    #Calculate royalty payments from statements and contracts

    def __init__(self):
        # Load the API key from Streamlit or .env
        api_key = (
            st.secrets.get("OPENAI_API_KEY")
            or os.getenv("OPENAI_API_KEY")
        )

        if not api_key or not api_key.startswith("sk-"):
            raise ValueError("❌ OpenAI API key missing. Please add it to Streamlit Secrets or .env.")

        # Pass key explicitly to avoid missing context
        self.contract_parser = MusicContractParser(api_key=api_key)
    
    def read_royalty_statement(self, excel_path: str, title_column: str = None, 
                              payable_column: str = None) -> Dict[str, float]:
        """
        Read streaming royalty statement and calculate total per song
        
        Args:
            excel_path: Path to the Excel royalty statement
            title_column: Name of column containing song titles (auto-detects if None)
            payable_column: Name of column containing net payable amounts (auto-detects if None)
            
        Returns:
            Dictionary mapping song titles to total net payable amounts
        """
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            sheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
            print(f"📋 Extracted headers: {headers}")
                
            
            # Auto-detect columns if not specified
            if title_column is None:
                title_column = self._find_title_column(headers)
            else:
                title_column = title_column.lower()
            
            if payable_column is None:
                payable_column = self._find_payable_column(headers)
            else:
                payable_column = payable_column.lower()
            
            # Find column indices
            try:
                title_idx = headers.index(title_column)
                payable_idx = headers.index(payable_column)
            except ValueError as e:
                raise ValueError(f"Could not find required columns. Available columns: {headers}")
            
            # Read data and sum by song title
            song_totals = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[title_idx] and row[payable_idx] is not None:
                    title = str(row[title_idx]).strip()
                    try:
                        amount = float(row[payable_idx])
                        song_totals[title] = song_totals.get(title, 0.0) + amount
                    except (ValueError, TypeError):
                        continue
            
            workbook.close()
            
            print(f"✓ Read {len(song_totals)} unique songs from royalty statement")
            return song_totals
            
        except Exception as e:
            raise Exception(f"Error reading royalty statement: {str(e)}")
    
    def _find_title_column(self, headers: List[str]) -> str:
        """Auto-detect the title column"""
        title_variations = ['release title', 'title', 'song title', 'track title', 
                           'song name', 'track name', 'release name', 'track']
        
        for header in headers:
            if any(var in header for var in title_variations):
                print(f"TITLE COLUMN: {header}")
                return header
        
        raise ValueError(f"Could not auto-detect title column. Please specify title_column parameter.")
    
    def _find_payable_column(self, headers: List[str]) -> str:
        """Auto-detect the net payable column"""
        # Prioritize more specific matches first
        priority_variations = [
            'net payable',
            'net payment', 
            'total payable',
            'net revenue',
            'net amount'
        ]
        
        # Check priority variations first (exact/specific matches)
        for header in headers:
            for var in priority_variations:
                if var == header or var in header:
                    print(f"PAYABLE COLUMN: {header}")
                    return header
        
        # Fallback to general variations if no priority match
        general_variations = ['payable', 'amount', 'earnings', 'payment']
        
        for header in headers:
            for var in general_variations:
                # Only match if it's not a "withheld" or "deduction" column
                if var in header and 'withheld' not in header and 'deduction' not in header and 'fee' not in header:
                    print(f"PAYABLE COLUMN: {header}")
                    return header
        
        raise ValueError(f"Could not auto-detect payable column. Please specify payable_column parameter.")
    
    def calculate_payments(self, contract_path: str, statement_path: str, 
                          title_column: str = None, payable_column: str = None) -> List[RoyaltyPayment]:
        """
        Calculate payments for all parties based on contract and royalty statement
        
        Args:
            contract_path: Path to the contract file (PDF or Excel)
            statement_path: Path to the royalty statement (Excel)
            title_column: Optional - Name of title column in statement
            payable_column: Optional - Name of payable column in statement
            
        Returns:
            List of RoyaltyPayment objects with calculated amounts
        """
        print("\n" + "="*80)
        print("ROYALTY PAYMENT CALCULATION")
        print("="*80)
        
        # Parse the contract
        print("\nStep 1: Parsing contract...")
        contract_data = self.contract_parser.parse_contract(contract_path)
        
        # Read royalty statement
        print("\nStep 2: Reading royalty statement...")
        song_totals = self.read_royalty_statement(statement_path, title_column, payable_column)
        
        # Calculate payments
        print("\n💰 Step 3: Calculating payments...\n")
        payments = []
        
        # Get streaming royalty shares from contract
        streaming_shares = [
            share for share in contract_data.royalty_shares 
            if 'streaming' in share.royalty_type.lower()
        ]
        
        # For each work in the contract
        for work in contract_data.works:
            # Find matching song in statement (fuzzy match)
            matching_song, total_royalty = self._find_matching_song(work.title, song_totals)
            
            if matching_song:
                print(f"✓ Found '{work.title}' in statement: ${total_royalty:,.2f} total royalties")
                
                # Calculate payment for each party with streaming shares
                for share in streaming_shares:
                    amount_to_pay = total_royalty * (share.percentage / 100.0)
                    
                    # Find party details
                    party = next((p for p in contract_data.parties 
                                if p.name.lower() == share.party_name.lower()), None)
                    role = party.role if party else "Unknown"
                    
                    payment = RoyaltyPayment(
                        song_title=work.title,
                        party_name=share.party_name,
                        role=role,
                        royalty_type=share.royalty_type,
                        percentage=share.percentage,
                        total_royalty=total_royalty,
                        amount_to_pay=amount_to_pay
                    )
                    payments.append(payment)
                    
                    print(f"  → {share.party_name} ({role}): {share.percentage}% = ${amount_to_pay:,.2f}")
            else:
                print(f"⚠ Warning: '{work.title}' not found in royalty statement")
        
        return payments
    

    def merge_contracts(self, contracts: list[ContractData]) -> ContractData:
        """
        Merge multiple ContractData objects for the same work.
        Combines parties, works, and royalty shares safely.
        Ensures no contributor is lost due to name formatting or overlap.
        """

        merged_parties = []
        merged_works = []
        merged_royalty_shares = []
        summaries = []

        # Helper function to normalize names
        def normalize_name(name: str) -> str:
            if not name:
                return ""
            # Lowercase, strip spaces and role annotations
            clean = re.sub(r"\(.*?\)", "", name).strip().lower()
            return clean

        seen_parties = set()
        seen_works = set()
        seen_shares = set()  # prevent duplicate (party, type, %) combos

        for contract in contracts:
            summaries.append(contract.contract_summary or "")

            # Merge parties
            for p in contract.parties:
                norm_name = normalize_name(p.name)
                if norm_name and norm_name not in seen_parties:
                    merged_parties.append(p)
                    seen_parties.add(norm_name)

            # Merge works
            for w in contract.works:
                norm_title = (w.title or "").strip().lower()
                if norm_title and norm_title not in seen_works:
                    merged_works.append(w)
                    seen_works.add(norm_title)

            # Merge royalty shares, but avoid duplicates
            for r in contract.royalty_shares:
                key = (
                    normalize_name(r.party_name),
                    r.royalty_type.lower().strip(),
                    round(float(r.percentage), 2)
                )
                if key not in seen_shares:
                    merged_royalty_shares.append(r)
                    seen_shares.add(key)

        merged_summary = "\n".join([s for s in summaries if s.strip()])

        print(f"✅ Merged {len(merged_parties)} parties, {len(merged_works)} works, and {len(merged_royalty_shares)} royalty entries.")

        return ContractData(
            parties=merged_parties,
            works=merged_works,
            royalty_shares=merged_royalty_shares,
            contract_summary=merged_summary
        )

    def calculate_payments_from_data(
        self,
        contract_data,  # type: ContractData
        statement_path: str,
        title_column: str = None,
        payable_column: str = None
    ):
        """
        Calculate payments for all parties based on a parsed ContractData object
        and a royalty statement file.

        Args:
            contract_data: Parsed and possibly merged ContractData object.
            statement_path: Path to the royalty statement (Excel).
            title_column: Optional name of title column in the statement.
            payable_column: Optional name of payable column in the statement.

        Returns:
            List of RoyaltyPayment objects with calculated amounts.
        """
        print("\n" + "="*80)
        print("ROYALTY PAYMENT CALCULATION (from parsed contract data)")
        print("="*80)

        # Step 1 — Validate contract data
        if not contract_data.works:
            raise ValueError("❌ No works found in the provided contract data.")
        if not contract_data.royalty_shares:
            raise ValueError("❌ No royalty share data found in the provided contract data.")

        # Step 2 — Read royalty statement
        print("\n📊 Reading royalty statement...")
        song_totals = self.read_royalty_statement(statement_path, title_column, payable_column)
        if not song_totals:
            raise ValueError("❌ No songs found in the royalty statement.")

        # Step 3 — Filter for streaming royalties
        streaming_shares = [
            share for share in contract_data.royalty_shares
            if "streaming" in share.royalty_type.lower()
        ]

        if not streaming_shares:
            print("⚠️ No streaming royalty shares found in this contract data.")
            return []

        # Step 4 — Calculate payments
        print("\n💰 Calculating payments...\n")
        payments = []

        for work in contract_data.works:
            # Find matching song title in the royalty statement (fuzzy match)
            matching_song, total_royalty = self._find_matching_song(work.title, song_totals)

            if matching_song:
                print(f"✓ Found '{work.title}' in statement: ${total_royalty:,.2f} total royalties")

                for share in streaming_shares:
                    amount_to_pay = total_royalty * (share.percentage / 100.0)

                    # Try to find party details
                    party = next(
                        (p for p in contract_data.parties if p.name.lower() == share.party_name.lower()),
                        None
                    )
                    role = party.role if party else "Unknown"

                    payment = RoyaltyPayment(
                        song_title=work.title,
                        party_name=share.party_name,
                        role=role,
                        royalty_type=share.royalty_type,
                        percentage=share.percentage,
                        total_royalty=total_royalty,
                        amount_to_pay=amount_to_pay
                    )
                    payments.append(payment)

                    print(f"  → {share.party_name} ({role}): {share.percentage}% = ${amount_to_pay:,.2f}")

            else:
                print(f"⚠️ Warning: '{work.title}' not found in royalty statement")

        # Step 5 — Return final payments
        print(f"\n✅ Calculated {len(payments)} payments total")
        return payments

    def calculate_payments_from_contracts(
        self,
        contract_paths: list[str],
        statement_path: str, 
    ):
        """
        Parse multiple contracts, merge their data, and calculate payments.

        Args:
            contract_paths: List of paths to contract files (PDF or Excel)
            statement_path: Path to the royalty statement file
            title_column: Optional column name for song titles in statement
            payable_column: Optional column name for payable amounts

        Returns:
            List of RoyaltyPayment objects with combined results
        """
        print(f"\nParsing {len(contract_paths)} contracts...")

        # Step 1 — Parse each contract using GPT-based parser
        all_contracts_data = []
        for path in contract_paths:
            try:
                data = self.contract_parser.parse_contract(path)
                all_contracts_data.append(data)
            except Exception as e:
                print(f"Warning: Failed to parse {path}: {e}")

        if not all_contracts_data:
            raise ValueError(" No valid contracts could be parsed. Please check your uploads.")

        # Step 2 — Merge all contract data
        merged_data = self.merge_contracts(all_contracts_data)
        print(f"Merged {len(all_contracts_data)} contracts into one combined dataset")

        # Step 3 — Use the merged contract data to calculate payouts
        payments = self.calculate_payments_from_data(
            merged_data,
            statement_path,
        )

        print(f"Calculated payments for {len(payments)} contributors")
        return payments
    
    def _find_matching_song(self, song_title: str, song_totals: Dict[str, float]) -> tuple:
        """
        Find matching song in royalty statement (with fuzzy matching)
        
        Returns:
            Tuple of (matched_title, total_amount) or (None, 0.0) if not found
        """
        song_title_lower = song_title.lower().strip()
        
        # Try exact match first
        for title, amount in song_totals.items():
            if title.lower().strip() == song_title_lower:
                return (title, amount)
        
        # Try partial match
        for title, amount in song_totals.items():
            if song_title_lower in title.lower() or title.lower() in song_title_lower:
                return (title, amount)
        
        return (None, 0.0)
    
    def save_payments_to_excel(self, payments: List[RoyaltyPayment], output_path: str):
        """
        Save calculated payments to an Excel file
        
        Args:
            payments: List of RoyaltyPayment objects
            output_path: Path to save the Excel file
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Royalty Payments"
        
        # Headers
        headers = ["Song Title", "Payee Name", "Role", "Royalty Type", 
                  "Share %", "Total Royalty", "Amount to Pay"]
        sheet.append(headers)
        
        # Make headers bold
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Add data
        for payment in payments:
            sheet.append([
                payment.song_title,
                payment.party_name,
                payment.role,
                payment.royalty_type,
                f"{payment.percentage}%",
                f"${payment.total_royalty:,.2f}",
                f"${payment.amount_to_pay:,.2f}"
            ])
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output_path)
        print(f"\nPayment breakdown saved to {output_path}")
    
    def save_payments_to_json(self, payments: List[RoyaltyPayment], output_path: str):
        """
        Save calculated payments to a JSON file
        
        Args:
            payments: List of RoyaltyPayment objects
            output_path: Path to save the JSON file
        """
        import json
        
        data = [asdict(payment) for payment in payments]
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"Payment data saved to {output_path}")
    
    def print_payment_summary(self, payments: List[RoyaltyPayment]):
        """Print a formatted summary of payments"""
        
        print("\n" + "="*80)
        print("PAYMENT SUMMARY")
        print("="*80)
        
        if not payments:
            print("\n⚠ No payments calculated")
            return
        
        # Group by payee
        payee_totals = {}
        for payment in payments:
            if payment.party_name not in payee_totals:
                payee_totals[payment.party_name] = {
                    'role': payment.role,
                    'total': 0.0,
                    'details': []
                }
            payee_totals[payment.party_name]['total'] += payment.amount_to_pay
            payee_totals[payment.party_name]['details'].append(payment)
        
        # Print summary for each payee
        for payee, data in payee_totals.items():
            print(f"\n👤 {payee} ({data['role']})")
            print(f"   Total Payment: ${data['total']:,.2f}")
            print(f"   Breakdown:")
            for detail in data['details']:
                print(f"      • {detail.song_title}: {detail.percentage}% of ${detail.total_royalty:,.2f} = ${detail.amount_to_pay:,.2f}")
        
        print(f"\n{'='*80}")
        print(f"GRAND TOTAL: ${sum(p.amount_to_pay for p in payments):,.2f}")
        print(f"{'='*80}\n")




def main():
    """Example usage of the RoyaltyCalculator"""
    
    calculator = RoyaltyCalculator()
    
    # Replace these with your actual file paths
    contract_path = "/Users/kenjiniyokindi/Documents/GREENBOX ANALYTICS/Personal Projects/Music Data Projects/Royalty automation/Music Contract.pdf"
    statement_path = "/Users/kenjiniyokindi/Documents/GREENBOX ANALYTICS/Personal Projects/Music Data Projects/Royalty automation/Royalty Statement.xlsx"
    
    try:
        # Calculate payments
        payments = calculator.calculate_payments(contract_path, statement_path)
        
        # Print summary to console
        calculator.print_payment_summary(payments)
        
        # Save to Excel
        calculator.save_payments_to_excel(payments, "royalty_payments.xlsx")
        
        # Save to JSON (optional)
        calculator.save_payments_to_json(payments, "royalty_payments.json")
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()


